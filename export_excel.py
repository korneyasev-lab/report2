import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import os

def create_excel_report(report_name, form_name, month, year, answers):
    """Создает Excel документ с отчетом"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет"

    title_font = Font(name='Arial', size=14, bold=True)
    question_font = Font(name='Arial', size=11, bold=True)
    answer_font = Font(name='Arial', size=12, bold=True)
    comment_font = Font(name='Arial', size=10, italic=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.merge_cells('A1:B1')
    ws['A1'] = report_name
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    current_row = 3

    ws.column_dimensions['A'].width = 65
    ws.column_dimensions['B'].width = 10

    for answer in answers:
        ws[f'A{current_row}'] = answer['question_text']
        ws[f'A{current_row}'].font = question_font
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws[f'A{current_row}'].border = thin_border

        ws[f'B{current_row}'] = answer['answer_yes_no']
        ws[f'B{current_row}'].font = answer_font
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{current_row}'].border = thin_border

        current_row += 1

        if answer['comment']:
            ws.merge_cells(f'A{current_row}:B{current_row}')
            ws[f'A{current_row}'] = answer['comment']
            ws[f'A{current_row}'].font = comment_font
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws[f'A{current_row}'].border = thin_border
            current_row += 1

    current_row += 1
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'] = f"Дата создания отчета: {datetime.now().strftime('%d.%m.%Y')}"
    ws[f'A{current_row}'].font = Font(name='Arial', size=11)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')

    current_row += 2
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'] = "Подпись: _________________________"
    ws[f'A{current_row}'].font = Font(name='Arial', size=11)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.print_options.horizontalCentered = True

    if not os.path.exists("отчеты"):
        os.makedirs("отчеты")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"отчеты/{report_name}_{timestamp}.xlsx"

    wb.save(filename)
    return filename